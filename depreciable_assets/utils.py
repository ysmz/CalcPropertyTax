import openpyxl, copy
from openpyxl.utils import get_column_letter

jpToJpNum = {"明":1, "大":2, "昭":3, "平":4, "令":5}

def convert_jp_to_jp_num(jp):
    return jpToJpNum[jp] if jp in jpToJpNum else 0
    
def range_copy_cell(sheet, min_col, min_row, max_col, max_row, shift_col, shift_row):
    merged_cells = copy.deepcopy(sheet.merged_cells);
    for merged_cell in merged_cells :
        if merged_cell.min_row >= min_row \
            and merged_cell.min_col >= min_col \
            and merged_cell.max_row <= max_row \
            and merged_cell.max_col <= max_col :
                sheet.unmerge_cells(merged_cell.coord);

    for col in range( min_col, max_col+1):
        for row in range( min_row, max_row+1):
            fmt = "{min_col}{min_row}"
            copySrcCoord = fmt.format(
                min_col = get_column_letter(col),
                min_row = row );

            copyDstCoord = fmt.format(
                min_col = get_column_letter(col + shift_col),
                min_row = row + shift_row );
            if type(sheet[copySrcCoord]) != 'MergedCell' :
                sheet[copyDstCoord].value = sheet[copySrcCoord].value;
                if sheet[copySrcCoord].has_style :
                    sheet[copyDstCoord]._style = sheet[copySrcCoord]._style;

    for merged_cell in merged_cells :
        if merged_cell.min_row >= min_row \
            and merged_cell.min_col >= min_col \
            and merged_cell.max_row <= max_row \
            and merged_cell.max_col <= max_col :
                sheet.merge_cells(merged_cell.coord);
                newCellRange = copy.copy(merged_cell);
                newCellRange.shift(shift_col, shift_row);
                sheet.merge_cells(newCellRange.coord);
    return 0;
    